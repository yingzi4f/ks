#!/usr/bin/env python3
"""将保安员考试试题 XLS 导入 SQLite，并导出前端 JSON。"""

from __future__ import annotations

import argparse
import json
import re
import sqlite3
import sys
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

OPTION_SPLIT = re.compile(
    r"(?m)^([A-Ha-h])[:：][ \t]*(.*?)(?=\n[A-Ha-h][:：]|\Z)",
    re.S,
)
JUDGE_TEXTS = {"正确", "错误", "对", "错", "是", "否"}


def load_sheet(xls_path: Path):
    raw = xls_path.read_bytes()
    wb = load_workbook(BytesIO(raw), data_only=True, read_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = list(row) + [None, None, None]
        rows.append((i, cells[0], cells[1], cells[2]))
    wb.close()
    return rows


def parse_options(raw) -> list[dict]:
    if raw is None:
        return []
    text = str(raw).replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return []
    options = []
    for match in OPTION_SPLIT.finditer(text):
        key = match.group(1).upper()
        value = re.sub(r"[ \t]*\n[ \t]*", "", match.group(2)).strip()
        options.append({"key": key, "text": value})
    if options:
        return options
    # 兜底：整段当作未知选项
    return [{"key": "A", "text": text}]


def parse_answer_keys(raw, options: list[dict]) -> list[str]:
    if raw is None:
        return []
    text = (
        str(raw)
        .strip()
        .upper()
        .replace("，", ",")
        .replace("、", ",")
        .replace(" ", "")
        .replace(";", ",")
        .replace("；", ",")
    )
    if not text:
        return []
    keys = [part for part in text.split(",") if part]
    valid = {opt["key"] for opt in options}
    # 保留原始答案字母，即使源数据缺选项
    return [k for k in keys if re.fullmatch(r"[A-H]", k) or k in valid]


def classify(options: list[dict], answer_keys: list[str]) -> str:
    texts = {opt["text"].strip("。．. ") for opt in options}
    if len(options) <= 2 and texts and texts.issubset(JUDGE_TEXTS):
        return "判断"
    if len(answer_keys) > 1:
        return "多选"
    return "单选"


def build_answer_text(options: list[dict], answer_keys: list[str]) -> str:
    mapping = {opt["key"]: opt["text"] for opt in options}
    parts = []
    for key in answer_keys:
        text = mapping.get(key, "")
        parts.append(f"{key}. {text}".rstrip() if text else key)
    return "；".join(parts)


def extract_questions(rows: list[tuple]) -> list[dict]:
    header = None
    questions = []
    seq = 0
    for excel_row, stem, options_raw, answer_raw in rows:
        if stem is None:
            continue
        stem_text = str(stem).strip()
        if not stem_text:
            continue
        if header is None and stem_text in {"题干", "题目", "问题"}:
            header = (stem_text, str(options_raw or ""), str(answer_raw or ""))
            continue
        options = parse_options(options_raw)
        answer_keys = parse_answer_keys(answer_raw, options)
        seq += 1
        qtype = classify(options, answer_keys)
        questions.append(
            {
                "id": seq,
                "excel_row": excel_row,
                "stem": stem_text,
                "options": options,
                "options_raw": "" if options_raw is None else str(options_raw),
                "answer_keys": answer_keys,
                "answer_text": build_answer_text(options, answer_keys),
                "qtype": qtype,
            }
        )
    return questions


def init_db(db_path: Path) -> sqlite3.Connection:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    if db_path.exists():
        db_path.unlink()
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode = DELETE")
    conn.execute("PRAGMA foreign_keys = ON")
    conn.executescript(
        """
        CREATE TABLE questions (
            id INTEGER PRIMARY KEY,
            excel_row INTEGER NOT NULL,
            stem TEXT NOT NULL,
            options_raw TEXT NOT NULL,
            options_json TEXT NOT NULL,
            answer_keys TEXT NOT NULL,
            answer_text TEXT NOT NULL,
            qtype TEXT NOT NULL
        );

        CREATE VIRTUAL TABLE questions_fts USING fts5(
            stem,
            options_raw,
            answer_text,
            qtype,
            content='questions',
            content_rowid='id',
            tokenize='unicode61 remove_diacritics 2'
        );

        CREATE TRIGGER questions_ai AFTER INSERT ON questions BEGIN
            INSERT INTO questions_fts(rowid, stem, options_raw, answer_text, qtype)
            VALUES (new.id, new.stem, new.options_raw, new.answer_text, new.qtype);
        END;
        """
    )
    return conn


def insert_questions(conn: sqlite3.Connection, questions: list[dict]) -> None:
    conn.executemany(
        """
        INSERT INTO questions (
            id, excel_row, stem, options_raw, options_json,
            answer_keys, answer_text, qtype
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                q["id"],
                q["excel_row"],
                q["stem"],
                q["options_raw"],
                json.dumps(q["options"], ensure_ascii=False),
                ",".join(q["answer_keys"]),
                q["answer_text"],
                q["qtype"],
            )
            for q in questions
        ],
    )
    conn.commit()


def export_json(json_path: Path, questions: list[dict], xls_path: Path) -> None:
    json_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "meta": {
            "title": "保安员考试试题",
            "source": xls_path.name,
            "count": len(questions),
            "imported_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        },
        "questions": [
            {
                "id": q["id"],
                "stem": q["stem"],
                "options": q["options"],
                "answer_keys": q["answer_keys"],
                "answer_text": q["answer_text"],
                "qtype": q["qtype"],
            }
            for q in questions
        ],
    }
    json_path.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="导入保安员考试试题 XLS")
    parser.add_argument("--xls", required=True, type=Path)
    parser.add_argument("--db", required=True, type=Path)
    parser.add_argument("--json", required=True, type=Path)
    args = parser.parse_args()

    if not args.xls.exists():
        print(f"找不到试题文件: {args.xls}", file=sys.stderr)
        return 1

    rows = load_sheet(args.xls)
    questions = extract_questions(rows)
    if not questions:
        print("没有解析到题目", file=sys.stderr)
        return 1

    conn = init_db(args.db)
    try:
        insert_questions(conn, questions)
        total = conn.execute("SELECT COUNT(*) FROM questions").fetchone()[0]
        types = dict(
            conn.execute(
                "SELECT qtype, COUNT(*) FROM questions GROUP BY qtype"
            ).fetchall()
        )
    finally:
        conn.close()

    export_json(args.json, questions, args.xls)
    print(f"导入完成: {total} 题")
    print(f"题型分布: {types}")
    print(f"SQLite: {args.db}")
    print(f"JSON:   {args.json}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
